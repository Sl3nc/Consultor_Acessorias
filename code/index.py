from pathlib import Path
import os
import sys
import traceback
from unidecode import unidecode
from time import sleep
from datetime import datetime, date

import pandas as pd
from pandas.errors import ParserError
from openpyxl import Workbook, load_workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter

from selenium.webdriver.remote.webelement import WebElement
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PySide6.QtWidgets import (
    QMainWindow, QApplication, QWidget, QLabel, QVBoxLayout,QPushButton, QLineEdit
)
from PySide6.QtGui import QPixmap, QIcon, QMovie
from PySide6.QtCore import QThread, QObject, Signal, QSize, QDate
from src.window_acessorias import Ui_MainWindow

from tkinter.messagebox import askyesno, showerror, showinfo
from tkinter.filedialog import askopenfilename, asksaveasfilename

import json
from dotenv import load_dotenv

from locale import setlocale, LC_ALL

setlocale(LC_ALL, 'pt_BR.UTF-8')

load_dotenv(Path(__file__).parent / 'src' / 'env' / '.env')

class Obrigacao:
    """
    Classe responsável por armazenar e manipular as obrigações de uma empresa,
    bem como suas informações básicas (nome e CNPJ).
    """
    def __init__(self, interesses: list) -> None:
        """
        Inicializa a Obrigacao com os interesses (tipos de obrigações) a serem acompanhados.
        """
        self.interesses = {key: list() for key in interesses}

        self.infos_empresa = {
            'Nome': list(),
            'CNPJ': list()
        }

        # self.obrigacao = {
        #     'GUIA FGTS DIGITAL': list(),
        #     'Pro labore': list(),
        #     'RESUMO FOLHA DE PAGAMENTO': list(),
        #     'BOLETO - HONORÁRIO CONTÁBIL': list(),
        # }
        pass

    def add_dados(self, dict_obriacoes: dict[str,str]):
        """
        Adiciona o status das obrigações conforme os dados extraídos do sistema.
        """
        for interesse, lista in self.interesses.items():
            lista.append('Pendente')
            for key, situacao in dict_obriacoes.items():
                # print(f'\nkey - {key}')
                # print(f'interesse - {interesse}\n')
                if interesse in key:
                    if 'Ent.' in situacao:
                        lista.pop()
                        lista.append(f'Enviado: {situacao[25:33]}')
                    break

    def add_empresa(self, infos_emp: list[str]):
        """
        Adiciona informações da empresa (nome e CNPJ) à Obrigacao.
        """
        for index, listas in enumerate(self.infos_empresa.values()):
            listas.append(infos_emp[index])

    def result(self):
        """
        Retorna um DataFrame pandas com as informações da empresa e obrigações.
        """
        return pd.DataFrame(self.infos_empresa | self.interesses)

class Matriz(QObject):
    """
    Classe responsável por manipular a matriz de empresas, validando e lendo os dados do arquivo Excel.
    """
    fim = Signal(dict)
    qnt_empresas = Signal(int)

    def __init__(self) -> None:
        super().__init__()
        self.tipos_validos = 'lsx'
        self.caminho = None
        self.planilha_excecoes = ['PORCENTAGEM', 'PERCENTUAIS']
        pass

    def inserir(self) -> str | None:
        """
        Abre um diálogo para o usuário selecionar o arquivo da matriz e valida o arquivo selecionado.
        """
        try:
            caminho = askopenfilename()
            if caminho == '':
                return None
            self.caminho = self.__validar_entrada(caminho)
            return self.caminho[self.caminho.rfind('/') +1:]

        except PermissionError:
            showerror(title='Aviso', message= 'O arquivo selecionado apresenta-se em aberto em outra janela, favor fecha-la')
            return None

        except FileExistsError:
            showerror(title='Aviso', message= 'O arquivo selecionado já apresenta uma versão sem acento, favor usar tal versão ou apagar uma delas')
            return None

        except Exception as error:
            showerror(title='Aviso', message= error)
            return None

    def __validar_entrada(self, caminho: str) -> str:
        """
        Valida o formato do arquivo e remove acentos do nome, se necessário.
        """
        if caminho[len(caminho) -3 :] != self.tipos_validos:
            ultima_barra = caminho.rfind('/')
            raise Exception(
                f'Formato inválido do arquivo: {caminho[ultima_barra+1:]}'
            )
        
        caminho_uni = unidecode(caminho)
        if caminho != caminho_uni:
            caminho = self.__renomear(caminho_uni)

        return caminho

    def __renomear(self, caminho, caminho_uni) -> str:
        """
        Renomeia o arquivo para remover acentos do nome.
        """
        os.renames(caminho, caminho_uni)
        return caminho
    
    def envio_invalido(self) -> bool:
        """
        Verifica se o caminho do arquivo foi definido corretamente.
        """
        return True if self.caminho == None else False

    def ler(self) -> dict[str,list[int]]:
        """
        Lê os dados das empresas da matriz a partir do arquivo Excel.
        """
        result = {}
        nomes_planilha = load_workbook(self.caminho).sheetnames
        for i in self.planilha_excecoes:
            if i in nomes_planilha:
                nomes_planilha.remove(i)

        count = 0
        for nome in nomes_planilha:
            result[nome] = [int(i) for i in pd.read_excel(self.caminho, usecols='A', header= None, sheet_name= nome).dropna().iloc[:, 0] if str(i)[0].isnumeric() == True]
            count = len(result[nome]) + count
        self.qnt_empresas.emit(count)
        self.fim.emit(result)

class Relatorio:
    """
    Classe responsável por gerar e salvar o relatório final em Excel, com os dados das obrigações das empresas.
    """
    TITULO = 'Relatório'

    def __init__(self) -> None:
        self.linha_cabecalho = 2
        self.linha_conteudo = 3
        self.espacos_tabela = [50,30,30,30,30,30,30,30]
        pass

    def nomear(self) -> str:
        """
        Solicita ao usuário o nome do arquivo para salvar o relatório.
        """
        nome_arq = asksaveasfilename(title='Favor nomear o arquivo que será salvo', filetypes=((".xlsx","*.xlsx"),))

        if nome_arq == '':
            if askyesno(title='Aviso', message= 'Deseja cancelar esta operação?') == True:
                raise Exception ('Operação cancelada!')
            else:
                return self.nomear()
            
        return nome_arq  + '.xlsx' 
    
    def alterar(self, data: dict[str, Obrigacao]) -> None:
        """
        Cria e salva o relatório Excel com os dados fornecidos.
        """
        wb = Workbook()
        del wb['Sheet']
        for key, obrigacao in data.items():
            df = obrigacao.result()
            if df.empty == False:
                ws = wb.create_sheet(key)
                self.width_ws(ws)

                self.fill_cabecalho(df, ws)
                self.fill_conteudo(df, ws)

        nome_arq = self.nomear()
        wb.save(nome_arq)
        
        showinfo(title='Aviso', message='Abrindo o arquivo gerado!')
        os.startfile(nome_arq)

    def fill_conteudo(self, conteudo: pd.DataFrame, ws):
        """
        Preenche o conteúdo do relatório na planilha.
        """
        for index_linha, row in conteudo.iterrows():
            for index_coluna, valor in enumerate(row, 1):
                ws.cell(index_linha + self.linha_conteudo, index_coluna).value = valor

    def fill_cabecalho(self, conteudo: pd.DataFrame, ws):
        """
        Preenche o cabeçalho do relatório na planilha.
        """
        for index_coluna, column in enumerate(conteudo.columns, 1):
            ws.cell(self.linha_cabecalho, index_coluna).value = CellRichText(
                TextBlock(InlineFont(b=True), column)
            )

    def width_ws(self, ws):
        """
        Ajusta a largura das colunas da planilha.
        """
        for index, valor in enumerate(self.espacos_tabela, 1):
            ws.column_dimensions[get_column_letter(index)].width = valor

class Acessorias:
    """
    Classe responsável por interagir com o sistema web Acessorias.com via Selenium,
    realizando login, pesquisa de entregas e extração de dados das empresas.
    """
    CHROME_DRIVER_PATH = Path(__file__).parent/'src'/'driver'/'chromedriver.exe'
    URL_MAIN = 'https://app.acessorias.com/sysmain.php'
    URL_ENTREGAS = 'https://app.acessorias.com/sysmain.php?m=3'
    URL_EMPRESA = 'https://app.acessorias.com/sysmain.php?m=4'

    INPUT_EMAIL = 'mailAC'
    INPUT_PASSWORD= 'passAC'
    BTN_ENTRAR = '#site-corpo > section.secao.secao-login > div > form > div.botoes > button'

    BTN_PESQUISA_ENTREGAS = 'searchEmp'
    BTN_PESQUISA_EMP = 'searchString'
    BTN_FILTRAR = 'btFilter'
    TABELA_ENTREGAS = 'divRelEntregas'

    COMPE_DE = 'EntCompDe'
    COMPE_PARA = 'EntCompAte'
    
    POSIC_NOME_EMP = '#divEmpZ_{0} > div.col-sm-5.col-xs-12.no-padding.aImage > span'

    POSIC_CNPJ_EMP = '#divEmpZ_{0} > div.col-sm-7.col-xs-12.no-padding.aImage > div:nth-child(1)'

    def __init__(self) -> None:
        self.class_status_entrega = "col-sm-3.col-xs-12.no-padding"
        self.class_nome_entrega = 'neg.brown'

        self.browser = self.make_chrome_browser(hide=True)
        self.browser.get(self.URL_MAIN)
        pass

    def make_chrome_browser(self,*options: str, hide: bool) -> webdriver.Chrome:
        """
        Inicializa o navegador Chrome com as opções fornecidas.
        """
        chrome_options = webdriver.ChromeOptions()

        if options is not None:
            for option in options:
                chrome_options.add_argument(option)

        chrome_service = Service(
            executable_path=str(self.CHROME_DRIVER_PATH),
        )

        browser = webdriver.Chrome(
            service=chrome_service,
            options=chrome_options
        )

        if hide == True:
            browser.set_window_position(-10000,0)

        return browser
    
    def login(self, usuario: str, senha: str):
        """
        Realiza login no sistema Acessorias.com.
        """
        self.browser.find_element(By.NAME, self.INPUT_EMAIL).send_keys(usuario)
        self.browser.find_element(By.NAME, self.INPUT_PASSWORD).send_keys(senha)

        self.browser.find_element(By.CSS_SELECTOR, self.BTN_ENTRAR).click()

    def pesquisar_entrega(self, num_empresa: str):
        """
        Pesquisa as entregas de uma empresa pelo número de domínio.
        """
        if self.browser.current_url != self.URL_ENTREGAS:
            self.browser.get(self.URL_ENTREGAS)

        for input in [self.BTN_PESQUISA_ENTREGAS]:
            self.browser.find_element(By.ID, input).clear()
            self.browser.find_element(By.ID, input)\
            .send_keys(str(num_empresa))

        self.browser.find_element(By.ID, self.BTN_FILTRAR).click()
        sleep(3)

        return self.extrair_dados(
            self.browser.find_element(By.ID, self.TABELA_ENTREGAS)
        )

    def set_competencia(self, competencia: str):
        """
        Define a competência (período) para filtrar as entregas.
        """
        self.browser.get(self.URL_ENTREGAS)

        for input in [self.COMPE_DE, self.COMPE_PARA]:
            self.browser.find_element(By.ID, input).clear()
            self.browser.find_element(By.ID, input)\
            .send_keys(competencia)

    def extrair_dados(self, tabela: WebElement):
        """
        Extrai os dados de entregas da tabela HTML.
        """
        result = {}

        nome_competencia = [
            x.text for x in tabela.find_elements(By.CLASS_NAME, self.class_nome_entrega)
        ]
        status = [
            x.text for x in tabela.find_elements(By.CLASS_NAME, self.class_status_entrega)
        ]
        count = 0
        for i in range(0, len(nome_competencia), 2):
            juncao = f'{nome_competencia[i]} {nome_competencia[i + 1]}'
            result[juncao] = status[count]
            count = count + 1
        
        return result

    def pesquisar_empresa(self, num_empresa: str):
        """
        Pesquisa e retorna informações da empresa (nome e CNPJ).
        """
        try:
            if self.browser.current_url != self.URL_EMPRESA:
                self.browser.get(self.URL_EMPRESA)

            self.browser.find_element(By.ID, self.BTN_PESQUISA_EMP).clear()

            self.browser.find_element(By.ID, self.BTN_PESQUISA_EMP)\
                .send_keys(num_empresa)

            self.browser.find_element(By.ID, self.BTN_FILTRAR).click()
            sleep(3)

            nome_emp = self.browser.find_element(By.CSS_SELECTOR, self.POSIC_NOME_EMP.format(num_empresa)).text

            cnpj_emp = self.browser.find_element(By.CSS_SELECTOR, self.POSIC_CNPJ_EMP.format(num_empresa)).text

            return [
                nome_emp[:nome_emp.rfind('[') - 1],
                cnpj_emp[:18]
            ]
        except:
            return [
                'Empresa não encontrada',
                'Num. domínio: '+ str(num_empresa)
            ]


    def close(self):
        """
        Fecha o navegador.
        """
        self.browser.close()

class Wellington(QObject):
    """
    Classe que coordena o processamento das empresas, realizando a extração dos dados de obrigações e geração do relatório final.
    """
    progress = Signal(int)
    fim = Signal()

    def __init__(self, itens_matriz: dict[str,list[int]], competencia: str) -> None:
        super().__init__()
        self.info_matriz = itens_matriz
        self.competencia = competencia

        self.obrigacoes = {
            'DOMESTICO': Obrigacao(
                ['Recibo de Pagamento de Salário','GUIA E-SOCIAL - EMPREGADOR','FOLHA DE PONTO']
            ),
            'SO PRO LABORE': Obrigacao(
                ['RESUMO FOLHA DE PAGAMENTO','DARF DCTFWEB','RECIBO DCTFWEB','ESOCIAL']
            ),
            'FOLHA SIMPLES': Obrigacao(
                ['RESUMO FOLHA DE PAGAMENTO','DARF','RECIBO DCTFWEB','ESOCIAL','GUIA FGTS DIGITAL','RELATORIO FGTS DIGITAL']
            ),
            'FOLHA COM MOVIMENTO': Obrigacao(
                ['RESUMO FOLHA DE PAGAMENTO','DARF','RECIBO DCTFWEB','ESOCIAL','GUIA FGTS DIGITAL','RELATORIO FGTS DIGITAL']
            ),
            'SEM MOVIMENTO': Obrigacao(['DARF DCTFWEB','RECIBO DCTFWEB','ESOCIAL']),
        }

        self.credenciais = [
            os.getenv("LOGIN",""),
            os.getenv("SENHA",""),
        ]

        pass

    #TODO TRABALHAR
    def trabalhar(self) -> pd.DataFrame:
        """
        Executa o fluxo principal: login, pesquisa de obrigações, coleta de dados e geração do relatório.
        """
        try:
            acessorias = Acessorias()
            acessorias.login(self.credenciais[0], self.credenciais[1])
            sleep(5)

            #Implementar procura pela obrigação e adição de dados ao mesmo
            count = 0
            acessorias.set_competencia(self.competencia)
            for categoria_matriz, list_num in self.info_matriz.items():
                for categoria_obri, obrigacao in self.obrigacoes.items():
                    if categoria_matriz == categoria_obri:
                        for num in list_num:
                            obrigacao.add_dados(acessorias.pesquisar_entrega(str(num)))
                            obrigacao.add_empresa(acessorias.pesquisar_empresa(str(num)))

                            count = count + 1
                            self.progress.emit(count)
                        break

            acessorias.close()
            Relatorio().alterar(self.obrigacoes)
            self.fim.emit()
        except Exception:
            traceback.print_exc()
            acessorias.close()
            
class MainWindow(QMainWindow, Ui_MainWindow):
    """
    Classe principal da interface gráfica, responsável por gerenciar as interações do usuário, progresso e execução das tarefas.
    """
    MAX_PROGRESS = 100

    def __init__(self, parent = None):
        super().__init__(parent)
        self.setupUi(self)
        self.matriz = Matriz()

        self.setWindowIcon(QIcon(
            (Path(__file__).parent /'src'/'imgs'/'acessorias_icon.ico').__str__()))
        self.logo.setPixmap(QPixmap(
            (Path(__file__).parent /'src'/'imgs'/'acessorias_hori.png').__str__()))
        icon = QIcon()
        icon.addFile(
            (Path(__file__).parent /'src'/'imgs'/'upload-icon.png').__str__(),QSize(),
            QIcon.Mode.Normal,
            QIcon.State.Off
        )
        self.pushButton_upload.setIcon(icon)
        self.movie = QMovie(
            (Path(__file__).parent /'src'/'imgs'/'load.gif').__str__()
        )
        self.load_movie.setMovie(self.movie)
        data_atual = datetime.now()
        self.dateEdit_competencia.setDate(
            QDate(data_atual.year, data_atual.month - 1, data_atual.day)
        )

        self.pushButton_upload.clicked.connect(self.inserir_arquivo)
        self.pushButton_enviar.clicked.connect(self.ler_matriz)

    def inserir_arquivo(self):
        """
        Ação do botão para inserir o arquivo da matriz.
        """
        resp = self.matriz.inserir()

        if resp != None:
            self.pushButton_upload.setText(resp)
            self.pushButton_upload.setIcon(QPixmap(''))

    def ler_matriz(self):
        """
        Inicia a leitura da matriz de empresas em uma thread separada.
        """
        if self.matriz.envio_invalido():
            raise Exception('Favor anexar seu relatório de processos')
         
        self.load_title.setText('Carregando empresas da matriz...')
        self.progressBar.hide()
        self.exec_load(True)

        self._thread = QThread()

        self.matriz.moveToThread(self._thread)
        self._thread.started.connect(self.matriz.ler)
        self.matriz.fim.connect(self._thread.quit)
        self.matriz.fim.connect(self._thread.deleteLater)
        self.matriz.fim.connect(self.hard_work)
        self._thread.finished.connect(self.matriz.deleteLater)
        self.matriz.qnt_empresas.connect(self.set_progress)

        self._thread.start()  

    #TODO HARD_WORK
    def hard_work(self, itens_matriz: dict):
        """
        Inicia o processamento das empresas e obrigações em uma thread separada.
        """
        try:
            self.load_title.setText('Pesquisando empresas...')
            self.progressBar.show()

            # for key, value in itens_matriz.items():
            #     print(f'{key} - {value}')

            self.wellington = Wellington(
                itens_matriz, self.dateEdit_competencia.text()
            )
            self._thread = QThread()

            self.wellington.moveToThread(self._thread)
            self._thread.started.connect(self.wellington.trabalhar)
            self.wellington.fim.connect(self._thread.quit)
            self.wellington.fim.connect(self._thread.deleteLater)
            self.wellington.fim.connect(self.encerramento)
            self._thread.finished.connect(self.wellington.deleteLater)
            self.wellington.progress.connect(self.add_progress)

            self._thread.start()  

        except ParserError:
            self.exec_load(False)
            showerror(title='Aviso', message= 'Erro ao ler o arquivo, certifique-se de ter inserido o arquivo correto')
        except Exception as err:
            self.exec_load(False)
            traceback.print_exc()
            showerror('Aviso', err)

    def encerramento(self):
        """
        Finaliza o processamento e exibe mensagem de sucesso.
        """
        self.exec_load(False)
        self.statusbar.showMessage(
            f'Execução com êxito às {datetime.now().strftime('%H:%M:%S')}', 10)
        
    def set_progress(self, valor):
        """
        Define o coeficiente de progresso para a barra de progresso.
        """
        self.coeficiente_progresso = self.MAX_PROGRESS / valor

    def add_progress(self, valor):
        """
        Atualiza o valor da barra de progresso.
        """
        self.progressBar.setValue(self.coeficiente_progresso * valor)

    def exec_load(self, action: bool):
        """
        Controla a exibição da animação de carregamento.
        """
        if action == True:
            self.movie.start()
            self.stackedWidget.setCurrentIndex(1)
        else:
            self.movie.stop()
            self.stackedWidget.setCurrentIndex(0)

if __name__ == '__main__':
    """
    Ponto de entrada da aplicação. Inicializa a interface gráfica.
    """
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec()