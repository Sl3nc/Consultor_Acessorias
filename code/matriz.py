from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showerror

from PySide6.QtCore import Signal, QObject
from openpyxl import load_workbook
from unidecode import unidecode
from os import renames
import pandas as pd

class Matriz(QObject):
    """
    Classe responsável por manipular a matriz de empresas, validando e lendo os dados do arquivo Excel.
    """
    fim = Signal(dict)
    qnt_empresas = Signal(int)

    def __init__(self) -> None:
        super().__init__()
        self.tipos_validos = 'lsx'
        self.caminho = None
        self.planilha_excecoes = ['PORCENTAGEM', 'PERCENTUAIS']
        pass

    def inserir(self) -> str | None:
        """
        Abre um diálogo para o usuário selecionar o arquivo da matriz e valida o arquivo selecionado.
        """
        try:
            caminho = askopenfilename()
            if caminho == '':
                return None
            self.caminho = self.__validar_entrada(caminho)
            return self.caminho[self.caminho.rfind('/') +1:]

        except PermissionError:
            showerror(title='Aviso', message= 'O arquivo selecionado apresenta-se em aberto em outra janela, favor fecha-la')
            return None

        except FileExistsError:
            showerror(title='Aviso', message= 'O arquivo selecionado já apresenta uma versão sem acento, favor usar tal versão ou apagar uma delas')
            return None

        except Exception as error:
            showerror(title='Aviso', message= error)
            return None

    def __validar_entrada(self, caminho: str) -> str:
        """
        Valida o formato do arquivo e remove acentos do nome, se necessário.
        """
        if caminho[len(caminho) -3 :] != self.tipos_validos:
            ultima_barra = caminho.rfind('/')
            raise Exception(
                f'Formato inválido do arquivo: {caminho[ultima_barra+1:]}'
            )
        
        caminho_uni = unidecode(caminho)
        if caminho != caminho_uni:
            caminho = self.__renomear(caminho_uni)

        return caminho

    def __renomear(self, caminho, caminho_uni) -> str:
        """
        Renomeia o arquivo para remover acentos do nome.
        """
        renames(caminho, caminho_uni)
        return caminho
    
    def envio_invalido(self) -> bool:
        """
        Verifica se o caminho do arquivo foi definido corretamente.
        """
        return True if self.caminho == None else False

    def ler(self) -> dict[str,list[int]]:
        """
        Lê os dados das empresas da matriz a partir do arquivo Excel.
        """
        result = {}
        nomes_planilha = load_workbook(self.caminho).sheetnames
        for i in self.planilha_excecoes:
            if i in nomes_planilha:
                nomes_planilha.remove(i)

        count = 0
        for nome in nomes_planilha:
            result[nome] = [int(i) for i in pd.read_excel(self.caminho, usecols='A', header= None, sheet_name= nome).dropna().iloc[:, 0] if str(i)[0].isnumeric() == True]
            count = len(result[nome]) + count
        self.qnt_empresas.emit(count)
        self.fim.emit(result)