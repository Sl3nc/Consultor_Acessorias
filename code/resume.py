from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl import Workbook

from tkinter.messagebox import showinfo, askyesno
from tkinter.filedialog import asksaveasfilename
from obligation import Obrigacao
from os import startfile
import pandas as pd

class Relatorio:
    """
    Classe responsável por gerar e salvar o relatório final em Excel, com os dados das obrigações das empresas.
    """
    TITULO = 'Relatório'

    def __init__(self) -> None:
        self.linha_cabecalho = 2
        self.linha_conteudo = 3
        self.espacos_tabela = [50,30,30,30,30,30,30,30]
        pass

    def nomear(self) -> str:
        """
        Solicita ao usuário o nome do arquivo para salvar o relatório.
        """
        nome_arq = asksaveasfilename(title='Favor nomear o arquivo que será salvo', filetypes=((".xlsx","*.xlsx"),))

        if nome_arq == '':
            if askyesno(title='Aviso', message= 'Deseja cancelar esta operação?') == True:
                raise Exception ('Operação cancelada!')
            else:
                return self.nomear()
            
        return nome_arq  + '.xlsx' 
    
    def alterar(self, data: dict[str, Obrigacao]) -> None:
        """
        Cria e salva o relatório Excel com os dados fornecidos.
        """
        wb = Workbook()
        del wb['Sheet']
        for key, obrigacao in data.items():
            df = obrigacao.result()
            if df.empty == False:
                ws = wb.create_sheet(key)
                self.width_ws(ws)

                self.fill_cabecalho(df, ws)
                self.fill_conteudo(df, ws)

        nome_arq = self.nomear()
        wb.save(nome_arq)
        
        showinfo(title='Aviso', message='Abrindo o arquivo gerado!')
        startfile(nome_arq)

    def fill_conteudo(self, conteudo: pd.DataFrame, ws):
        """
        Preenche o conteúdo do relatório na planilha.
        """
        for index_linha, row in conteudo.iterrows():
            for index_coluna, valor in enumerate(row, 1):
                ws.cell(index_linha + self.linha_conteudo, index_coluna).value = valor

    def fill_cabecalho(self, conteudo: pd.DataFrame, ws):
        """
        Preenche o cabeçalho do relatório na planilha.
        """
        for index_coluna, column in enumerate(conteudo.columns, 1):
            ws.cell(self.linha_cabecalho, index_coluna).value = CellRichText(
                TextBlock(InlineFont(b=True), column)
            )

    def width_ws(self, ws):
        """
        Ajusta a largura das colunas da planilha.
        """
        for index, valor in enumerate(self.espacos_tabela, 1):
            ws.column_dimensions[get_column_letter(index)].width = valor